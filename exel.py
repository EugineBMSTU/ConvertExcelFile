import openpyxl
import xlwings as xw
dataBase = openpyxl.load_workbook(filename = 'dataBase.xlsx', data_only=True)
sheetdataBase = dataBase['list1']
for i in range(sheetdataBase.max_row):
    wb = xw.Book('coordConvert.xlsx')
    sht = wb.sheets['E,N Zne to Latitude & Longitude']
    sht.range('E4').value = sheetdataBase[('A' + str(i+1))].value
    sht.range('O4').value = sheetdataBase[('B' + str(i+1))].value
    sheetdataBase[('D' + str(i+1))].value = sht.range('C43').value
    sheetdataBase[('E' + str(i+1))].value = sht.range('D43').value
    sheetdataBase[('F' + str(i+1))].value = sht.range('E43').value
    sheetdataBase[('H' + str(i+1))].value = sht.range('M43').value
    sheetdataBase[('I' + str(i+1))].value = sht.range('N43').value
    sheetdataBase[('J' + str(i+1))].value = sht.range('O43').value

    if (i % 1000) == 0:
        dataBase.save('dataBase2.xlsx')

    print i*1.0/sheetdataBase.max_row*100.0, '%'
print 'ok'
